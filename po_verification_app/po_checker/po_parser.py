"""Wczytywanie pliku PO.

Plik PO ma nietypową, "wielowierszową" strukturę: nagłówek rozciąga się na
dwa wiersze, a każda pozycja towarowa zajmuje dwa kolejne wiersze (dane +
wiersz z rozbiciem na sztuki w kartonie). Dlatego zamiast pandas używamy
openpyxl i poruszamy się po komórkach tak samo, jak robił to oryginalny
kod w JS (tam też była to ręczna pętla po `worksheet` z surowego XLSX).

Indeksy wierszy/kolumn w tym module są 0-indeksowane, żeby 1:1 odpowiadały
logice oryginału - stąd `cell_val` doda +1 przy komunikacji z openpyxl.
"""
import re
from io import BytesIO
from typing import List, Optional, Tuple

from openpyxl import load_workbook

from .models import PORecord
from .helpers import js_number, is_truthy_number, round4


def cell_val(ws, r: int, c: int):
    """Wartość komórki, r i c 0-indeksowane."""
    return ws.cell(row=r + 1, column=c + 1).value


def find_col(ws, row1: int, row2: int, col_start: int, col_end: int, v1, v2) -> Optional[int]:
    """Szuka kolumny, w której dwa wiersze nagłówka mają dokładnie wartości v1/v2."""
    for c in range(col_start, col_end + 1):
        if cell_val(ws, row1, c) == v1 and cell_val(ws, row2, c) == v2:
            return c
    return None


def find_header_rows(ws, max_row: int) -> Optional[Tuple[int, int]]:
    """Znajduje dwuwierszowy nagłówek, szukając komórki 'PID' w kolumnie A."""
    limit = min(29, max_row)
    for r in range(0, limit + 1):
        if cell_val(ws, r, 0) == 'PID':
            return r - 1, r
    return None


def find_po_number_in_content(ws, max_scan_rows: int = 20) -> Optional[str]:
    """Szuka numeru PO wewnątrz pliku, np. w linijce 'PO #: 7588488   Dept: 451...'.

    To jest dużo pewniejsze niż wyciąganie numeru z nazwy pliku, bo ta linijka
    jest generowana automatycznie przez system źródłowy (FedBuy) i nie zależy
    od tego, jak ktoś zapisał/pobrał plik.
    """
    for r in range(0, max_scan_rows):
        val = cell_val(ws, r, 0)
        if val is None:
            continue
        m = re.search(r'PO\s*#:?\s*(\d+)', str(val), re.IGNORECASE)
        if m:
            return m.group(1)
    return None


def guess_po_number_from_filename(filename: str) -> Optional[str]:
    """Ostatnia deska ratunku, gdyby numeru PO nie dało się znaleźć w treści pliku."""
    m = re.search(r'PO[ _#]*(\d+)', filename, re.IGNORECASE)
    return m.group(1) if m else None


def parse_po_workbook(file_bytes: bytes, filename: str) -> dict:
    """Parsuje plik PO i zwraca słownik z metadanymi + listą PORecord.

    Konto (Store / .com) NIE jest wyciągane z tego pliku - system źródłowy
    (FedBuy) go tam nie zapisuje, więc użytkownik podaje je ręcznie w
    interfejsie aplikacji.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    max_row = ws.max_row - 1
    max_col = ws.max_column - 1

    header_rows = find_header_rows(ws, max_row)
    if header_rows is None:
        raise ValueError('Nie udało się znaleźć wiersza nagłówka z "PID" w tym pliku')
    row1, row2 = header_rows

    col_nrf = find_col(ws, row1, row2, 0, max_col, 'NRF', 'Color')
    col_desc = find_col(ws, row1, row2, 0, max_col, 'Color', 'Desc')
    col_entered = find_col(ws, row1, row2, 0, max_col, 'Entered', 'Qty')
    col_ext_cost = find_col(ws, row1, row2, 0, max_col, 'Ext', 'Cost $')
    col_ext_owned = find_col(ws, row1, row2, 0, max_col, 'Ext', 'Owned $')
    col_own_mu = find_col(ws, row1, row2, 0, max_col, 'Own', 'MU%')
    col_scale = find_col(ws, row1, row2, 0, max_col, 'Scale', 'Details')

    missing = []
    for label, val in [
        ('NRF Color', col_nrf), ('Color Desc', col_desc), ('Entered Qty', col_entered),
        ('Ext Cost $', col_ext_cost), ('Ext Owned $', col_ext_owned),
        ('Own MU%', col_own_mu), ('Scale Details', col_scale),
    ]:
        if val is None:
            missing.append(label)
    if missing:
        raise ValueError('Nie udało się znaleźć kolumn: ' + ', '.join(missing))

    records: List[PORecord] = []
    i = row2 + 1
    while i <= max_row:
        a = cell_val(ws, i, 0)
        if a is not None and a != 'Order Total':
            pid = str(a)
            j = i + 1
            while j <= max_row:
                aj = cell_val(ws, j, 0)
                if aj is not None:
                    break  # kolejny PID zaczyna się tutaj

                c = cell_val(ws, j, col_nrf)
                if c is None:
                    j += 1
                    continue

                d = cell_val(ws, j, col_desc)
                o = js_number(cell_val(ws, j, col_entered))
                v = js_number(cell_val(ws, j, col_ext_cost))
                y = js_number(cell_val(ws, j, col_ext_owned))
                z = js_number(cell_val(ws, j, col_own_mu))

                pack_row = j + 1
                vals = []
                for cc in range(col_scale + 1, col_entered):
                    raw = cell_val(ws, pack_row, cc)
                    if raw is not None and raw != '':
                        vals.append(js_number(raw))

                # jeśli ostatnia liczba to suma poprzednich, to jest to kolumna
                # "razem" - odrzucamy ją, zostają same rozmiary kartonu
                if len(vals) >= 2 and abs(sum(vals[:-1]) - vals[-1]) < 0.001:
                    pack_vals = vals[:-1]
                else:
                    pack_vals = vals
                pack_ratio = '-'.join(str(round(x)) for x in pack_vals)

                price = round4(v / o) if is_truthy_number(o) else None
                msrp = round4(y / o) if is_truthy_number(o) else None

                records.append(PORecord(
                    pid=pid,
                    pid6=pid[:6],
                    nrf_color=str(c).strip(),
                    color_desc=str(d).strip() if d else '',
                    pack_ratio=pack_ratio,
                    entered_qty=o,
                    ext_cost=v,
                    ext_owned=y,
                    own_mu=z,
                    price=price,
                    msrp=msrp,
                ))
                j += 2
            i = j
        else:
            i += 1

    po_number = find_po_number_in_content(ws) or guess_po_number_from_filename(filename)
    if not po_number:
        raise ValueError(
            'Nie udało się ustalić numeru PO ani z treści pliku, ani z nazwy pliku.'
        )

    return {
        'filename': filename,
        'po_number': po_number,
        'records': records,
    }
