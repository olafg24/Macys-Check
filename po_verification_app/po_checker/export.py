"""Budowanie pobieralnego raportu .xlsx — układ 1:1 ze wzorcem klienta
(PO_..._Verification.xlsx): najpierw wszystkie pola z PO, potem wszystkie
pola z systemu, potem Cancel Date/Reason, potem kolumny Match, na końcu Notes.
"""
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .matcher import FIELDS, HEADERS, compute_checks, po_sys_values

GREEN = 'FFC6EFCE'
RED = 'FFFFC7CE'
NAVY = 'FF1F4E78'

# Szerokości kolumn 1:1 ze wzorca dostarczonego przez klienta.
COLUMN_WIDTHS = [
    12, 10, 13, 13, 12, 12, 13, 10, 9, 9,     # 10x pola PO
    10, 15, 13, 12, 12, 13, 10, 9, 9,          # 9x pola Sys (bez PID)
    13, 10,                                    # Cancel Date / Reason
    10, 12, 13, 13, 12, 12, 13, 10, 9, 9,      # 10x kolumny Match
    45,                                         # Notes
]


def _fmt(v):
    return '' if v is None or v == '' else v


def build_excel_report(po_data: dict, results: List[dict]) -> BytesIO:
    """Zwraca bufor BytesIO gotowy do pobrania przez st.download_button."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'PO vs System Verification'

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    n_po_fields = len(FIELDS)          # 10
    n_sys_fields = len(FIELDS) - 1     # 9 (bez PID)
    match_start_col = n_po_fields + n_sys_fields + 2 + 1  # +2 = Cancel Date/Reason, +1 bo 1-indeksowane

    for rec in results:
        vals = po_sys_values(rec)

        po_values = [_fmt(vals[f][0]) for f in FIELDS]
        sys_values = [_fmt(vals[f][1]) for f in FIELDS if f != 'PID']
        cancel_reason = [
            rec.get('cancel_dates', '') if rec['sys_found'] else '',
            rec.get('sys_reason', '') if rec['sys_found'] else '',
        ]

        if rec['sys_found']:
            checks = compute_checks(rec)
            match_values = [checks[f] for f in FIELDS]
        else:
            match_values = ['' for _ in FIELDS]

        notes = rec.get('note') or ('' if rec['sys_found'] else 'No matching system line found')

        row_data = po_values + sys_values + cancel_reason + match_values + [notes]
        ws.append(row_data)
        row = ws[ws.max_row]

        if not rec['sys_found']:
            for cell in row:
                cell.fill = PatternFill('solid', fgColor=RED)
            continue

        for idx in range(n_po_fields):
            cell = row[match_start_col - 1 + idx]
            cell.fill = PatternFill('solid', fgColor=GREEN if match_values[idx] == 'MATCH' else RED)
            cell.alignment = Alignment(horizontal='center')

    for i, width in enumerate(COLUMN_WIDTHS):
        letter = ws.cell(row=1, column=i + 1).column_letter
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
